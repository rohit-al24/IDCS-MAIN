from fastapi import APIRouter, UploadFile, File, HTTPException
from typing import List
from openpyxl import load_workbook
from io import BytesIO
import re
import zipfile
import os
import xml.etree.ElementTree as ET
import posixpath

router = APIRouter()

@router.post("/upload-questions-excel/")
def upload_questions_excel(file: UploadFile = File(...)):
    try:
        data = file.file.read()
        wb = load_workbook(BytesIO(data), data_only=True)

        # preferred sheets order; process CO1-CO2 then CO3-CO4 then CO5 if present
        preferred_sheets = ['CO1-CO2', 'CO3-CO4', 'CO5']
        sheets_to_process = [s for s in preferred_sheets if s in wb.sheetnames]
        if not sheets_to_process:
            sheets_to_process = [wb.active.title]

        header_row_idx = 3
        required = ["Question Bank", "TYPE", "BTL Level", "Course Outcomes", "Marks", "Part"]

        try:
            from openpyxl.utils import coordinate_to_tuple
        except ImportError:
            def coordinate_to_tuple(cell):
                m = re.match(r"([A-Z]+)([0-9]+)", cell)
                if m:
                    col = 0
                    for c in m.group(1):
                        col = col * 26 + (ord(c) - ord('A') + 1)
                    return (int(m.group(2)), col)
                return (1, 1)

        def norm(s):
            return re.sub(r"\s+", "", s or "").replace("\n", "").replace("\r", "").lower()

        def cell_to_text(val):
            """Normalize Excel cell value to a string.
            Convert floats that are whole numbers to integers so '1.0' -> '1'.
            Return empty string for None or blank-like values.
            """
            if val is None:
                return ''
            # handle floats that are integers
            try:
                if isinstance(val, float):
                    if val.is_integer():
                        return str(int(val))
                    return str(val).strip()
                if isinstance(val, int):
                    return str(val)
            except Exception:
                pass
            # fallback
            try:
                return str(val).strip()
            except Exception:
                return ''

        all_questions = []

        for sheet_name in sheets_to_process:
            ws = wb[sheet_name]
            headers = [str(ws.cell(row=header_row_idx, column=c).value).strip() if ws.cell(row=header_row_idx, column=c).value is not None else '' for c in range(1, ws.max_column + 1)]
            norm_headers = [norm(h) for h in headers]
            header_map = {}
            used_indices = set()
            for col in required:
                norm_col = norm(col)
                found = False
                for idx, h in enumerate(norm_headers):
                    if idx in used_indices:
                        continue
                    if norm_col in h or h in norm_col:
                        header_map[col] = idx+1
                        used_indices.add(idx)
                        found = True
                        break
                if not found:
                    raise HTTPException(status_code=400, detail=f"Missing required column: {col} in sheet {sheet_name}. Found headers: {headers}")

            # image mapping per-sheet
            image_cell_map = {}
            image_row_map = {}
            imgs = list(getattr(ws, '_images', []))
            try:
                z = zipfile.ZipFile(BytesIO(data))
                try:
                    sheet_index = wb.sheetnames.index(ws.title) + 1
                except Exception:
                    sheet_index = 1
                sheet_path = f'xl/worksheets/sheet{sheet_index}.xml'
                if sheet_path in z.namelist():
                    sheet_xml = z.read(sheet_path)
                    st = ET.fromstring(sheet_xml)
                    drawing_elem = st.find('.//{http://schemas.openxmlformats.org/spreadsheetml/2006/main}drawing')
                    if drawing_elem is not None:
                        rId = drawing_elem.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
                        rels_path = f'xl/worksheets/_rels/sheet{sheet_index}.xml.rels'
                        if rels_path in z.namelist():
                            rels_doc = ET.fromstring(z.read(rels_path))
                            draw_target = None
                            for rel in rels_doc.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                                if rel.attrib.get('Id') == rId:
                                    draw_target = rel.attrib.get('Target')
                                    break
                            if draw_target:
                                if draw_target.startswith('../'):
                                    drawing_path = 'xl/' + draw_target.replace('../', '')
                                else:
                                    drawing_path = 'xl/' + draw_target.lstrip('./')
                                drawing_path = drawing_path.replace('\\', '/').replace('//', '/')
                                if drawing_path in z.namelist():
                                    drawing_xml = ET.fromstring(z.read(drawing_path))
                                    drawing_rels_path = os.path.dirname(drawing_path).rstrip('/') + '/_rels/' + os.path.basename(drawing_path) + '.rels'
                                    drawing_rels_path = drawing_rels_path.replace('\\', '/').replace('//', '/')
                                    rels_map = {}
                                    if drawing_rels_path in z.namelist():
                                        dr = ET.fromstring(z.read(drawing_rels_path))
                                        for rel in dr.findall('.//{http://schemas.openxmlformats.org/package/2006/relationships}Relationship'):
                                            Id = rel.attrib.get('Id')
                                            Target = rel.attrib.get('Target')
                                            if Id and Target:
                                                rel_base = os.path.dirname(drawing_path).replace('\\', '/').rstrip('/')
                                                combined = posixpath.normpath(posixpath.join(rel_base, Target)).replace('\\', '/')
                                                if not combined.startswith('xl/'):
                                                    combined = combined.lstrip('/')
                                                    combined = 'xl/' + combined
                                                rels_map[Id] = combined
                                    xdr_ns = '{http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing}'
                                    a_ns = '{http://schemas.openxmlformats.org/drawingml/2006/main}'
                                    anchors = drawing_xml.findall('.//'+xdr_ns+'twoCellAnchor') + drawing_xml.findall('.//'+xdr_ns+'oneCellAnchor')
                                    for anchor in anchors:
                                        frm = anchor.find('.//'+xdr_ns+'from')
                                        if frm is None:
                                            continue
                                        col_elem = frm.find('./'+xdr_ns+'col')
                                        row_elem = frm.find('./'+xdr_ns+'row')
                                        if col_elem is None or row_elem is None:
                                            continue
                                        try:
                                            a_col = int(col_elem.text) + 1
                                            a_row = int(row_elem.text) + 1
                                        except Exception:
                                            continue
                                        blip = anchor.find('.//'+xdr_ns+'pic//'+xdr_ns+'blipFill//'+a_ns+'blip')
                                        if blip is None:
                                            blip = anchor.find('.//'+xdr_ns+'blipFill//'+a_ns+'blip')
                                        if blip is None:
                                            continue
                                        embed = blip.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed')
                                        if not embed:
                                            continue
                                        media_target = rels_map.get(embed)
                                        if not media_target:
                                            continue
                                        media_path = str(media_target).replace('\\', '/').replace('//', '/')
                                        while media_path.startswith('./') or media_path.startswith('../'):
                                            if media_path.startswith('../'):
                                                media_path = media_path[3:]
                                            elif media_path.startswith('./'):
                                                media_path = media_path[2:]
                                        if not media_path.startswith('xl/'):
                                            media_path = 'xl/' + media_path.lstrip('/')
                                        media_path = media_path.replace('xl/xl/', 'xl/')
                                        if media_path in z.namelist():
                                            media_bytes = z.read(media_path)
                                            image_cell_map[(a_row, a_col)] = {'bytes': media_bytes, 'anchor_row': a_row, 'anchor_col': a_col}
                                            # map to nearest question row in this sheet
                                            target_row = a_row
                                            try:
                                                candidate_rows = []
                                                q_col_idx = header_map.get('Question Bank')
                                                if q_col_idx is not None:
                                                    for rr in range(header_row_idx + 1, ws.max_row + 1):
                                                        v = ws.cell(row=rr, column=q_col_idx).value
                                                        if v is not None and cell_to_text(v) != '':
                                                            candidate_rows.append(rr)
                                                if candidate_rows:
                                                    target_row = min(candidate_rows, key=lambda x: abs(x - a_row))
                                                else:
                                                    lo = max(header_row_idx + 1, a_row - 50)
                                                    hi = min(ws.max_row, a_row + 50)
                                                    for rr in range(lo, hi + 1):
                                                        v = ws.cell(row=rr, column=header_map.get('Question Bank')).value if header_map.get('Question Bank') is not None else None
                                                        if v is not None and cell_to_text(v) != '':
                                                            target_row = rr
                                                            break
                                            except Exception:
                                                target_row = a_row
                                            if target_row not in image_row_map:
                                                image_row_map[target_row] = {'anchor_row': a_row, 'anchor_col': a_col, 'bytes': media_bytes}
            except Exception as e:
                print(f"[Excel Debug] ZIP parse for sheet '{sheet_name}' failed: {e}")

            for img in imgs:
                anchor = getattr(img, 'anchor', None)
                if anchor is None:
                    continue
                if hasattr(anchor, '_from'):
                    a_row = anchor._from.row + 1
                    a_col = anchor._from.col + 1
                elif hasattr(anchor, 'cell'):
                    a_row, a_col = coordinate_to_tuple(anchor.cell)
                else:
                    continue
                q_col_idx = header_map.get('Question Bank')
                target_row = a_row
                if q_col_idx is not None:
                    try:
                        candidate_rows = [rr for rr in range(header_row_idx + 1, ws.max_row + 1) if cell_to_text(ws.cell(row=rr, column=q_col_idx).value) != '']
                        if candidate_rows:
                            target_row = min(candidate_rows, key=lambda x: abs(x - a_row))
                        else:
                            lo = max(header_row_idx + 1, a_row - 50)
                            hi = min(ws.max_row, a_row + 50)
                            for rr in range(lo, hi + 1):
                                if cell_to_text(ws.cell(row=rr, column=q_col_idx).value) != '':
                                    target_row = rr
                                    break
                    except Exception:
                        target_row = a_row
                image_cell_map[(a_row, a_col)] = img
                if target_row not in image_row_map:
                    image_row_map[target_row] = img

            # extract questions
            q_col_idx = header_map.get('Question Bank')
            for r in range(header_row_idx + 1, ws.max_row + 1):
                qtext = ''
                q_source_row = r
                q_source_col = q_col_idx
                if q_col_idx:
                    qcell_val = ws.cell(row=r, column=q_col_idx).value
                    qtext = cell_to_text(qcell_val)
                if not qtext and q_col_idx:
                    found = False
                    for d in range(1, 6):
                        rr = r - d
                        if rr > header_row_idx:
                            val = ws.cell(row=rr, column=q_col_idx).value
                            if val is not None and cell_to_text(val) != '':
                                qtext = cell_to_text(val)
                                q_source_row = rr
                                found = True
                                break
                    if not found:
                        for d in range(1, 6):
                            rr = r + d
                            if rr <= ws.max_row:
                                val = ws.cell(row=rr, column=q_col_idx).value
                                if val is not None and cell_to_text(val) != '':
                                    qtext = cell_to_text(val)
                                    q_source_row = rr
                                    break
                if not qtext:
                    continue

                def is_numeric_short(s: str) -> bool:
                    s2 = s.strip()
                    if not s2:
                        return True
                    if re.fullmatch(r"\d+", s2):
                        return True
                    if len(s2) <= 3:
                        return True
                    return False

                if is_numeric_short(qtext):
                    best = None
                    best_len = 0
                    for c in range(1, ws.max_column + 1):
                        if c == q_col_idx:
                            continue
                        v = ws.cell(row=r, column=c).value
                        if v is None:
                            continue
                        s = cell_to_text(v)
                        if s and not re.fullmatch(r"\d+", s) and len(s) > best_len:
                            best = (s, c)
                            best_len = len(s)
                    if best:
                        qtext, q_source_col = best[0], best[1]
                    else:
                        found = False
                        for d in range(1, 4):
                            for rr in (r - d, r + d):
                                if rr <= header_row_idx or rr > ws.max_row:
                                    continue
                                for c in range(1, ws.max_column + 1):
                                    v = ws.cell(row=rr, column=c).value
                                    if v is None:
                                        continue
                                    s = cell_to_text(v)
                                    if s and not re.fullmatch(r"\d+", s) and len(s) > 3:
                                        qtext = s
                                        q_source_row = rr
                                        q_source_col = c
                                        found = True
                                        break
                                if found:
                                    break
                            if found:
                                break

                type_raw = cell_to_text(ws.cell(row=r, column=header_map['TYPE']).value or '').lower()
                if type_raw == 'o':
                    qtype = 'objective'
                elif type_raw == 'd':
                    qtype = 'descriptive'
                elif type_raw == 'c':
                    qtype = 'Part_C'
                else:
                    continue

                btl_raw = ws.cell(row=r, column=header_map['BTL Level']).value
                btl = 2
                if btl_raw is not None:
                    btl_str = cell_to_text(btl_raw).upper()
                    btl_nums = re.findall(r'\d+', btl_str)
                    if btl_nums:
                        btl = int(max(btl_nums, key=int))

                marks_raw = ws.cell(row=r, column=header_map['Marks']).value
                try:
                    marks = int(marks_raw) if marks_raw is not None else 1
                except Exception:
                    marks = 1

                co_cell_value = ws.cell(row=r, column=header_map['Course Outcomes']).value
                co_raw = cell_to_text(co_cell_value).replace('\n',' ').replace('\r',' ').strip()
                digits_found = re.findall(r'([1-5])', co_raw)
                ordered_unique = []
                for d in digits_found:
                    if d not in ordered_unique:
                        ordered_unique.append(d)
                co_multi_numbers = ','.join(ordered_unique) if ordered_unique else ''
                co = None
                if ordered_unique:
                    co = f"CO{ordered_unique[0]}"
                if co is None and co_raw:
                    s_up = co_raw.upper()
                    m_single = re.search(r'CO\s*([1-5])', s_up)
                    if m_single:
                        co = f"CO{m_single.group(1)}"

                chapter = cell_to_text(ws.cell(row=r, column=header_map['Part']).value or '').strip() or None

                image_data = None
                img = None
                if q_source_row in image_row_map:
                    img = image_row_map[q_source_row]
                    mapped_row = q_source_row
                elif r in image_row_map:
                    img = image_row_map[r]
                    mapped_row = r
                else:
                    mapped_row = None
                    for c in range(1, ws.max_column + 1):
                        if (r, c) in image_cell_map:
                            img = image_cell_map[(r, c)]
                            mapped_row = r
                            break
                        if (q_source_row, c) in image_cell_map:
                            img = image_cell_map[(q_source_row, c)]
                            mapped_row = q_source_row
                            break

                image_anchor_row = None
                image_anchor_col = None
                image_present = False
                if img is not None:
                    import base64
                    img_bytes = None
                    if isinstance(img, dict):
                        image_anchor_row = img.get('anchor_row')
                        image_anchor_col = img.get('anchor_col')
                        img_bytes = img.get('bytes')
                    else:
                        anchor = getattr(img, 'anchor', None)
                        if anchor is not None:
                            if hasattr(anchor, '_from'):
                                image_anchor_row = anchor._from.row + 1
                                image_anchor_col = anchor._from.col + 1
                            elif hasattr(anchor, 'cell'):
                                rr, cc = coordinate_to_tuple(anchor.cell)
                                image_anchor_row = rr
                                image_anchor_col = cc
                        for attr in ('_data', 'image', 'ref', 'blob'):
                            v = getattr(img, attr, None)
                            if v:
                                if isinstance(v, bytes):
                                    img_bytes = v
                                    break
                                try:
                                    if hasattr(v, 'tobytes'):
                                        img_bytes = v.tobytes()
                                        break
                                    if hasattr(v, 'read'):
                                        img_bytes = v.read()
                                        break
                                except Exception:
                                    pass
                        if img_bytes is None:
                            try:
                                v = getattr(img, '_data', None)
                                if v is not None:
                                    if hasattr(v, 'tobytes'):
                                        img_bytes = v.tobytes()
                                    elif isinstance(v, bytes):
                                        img_bytes = v
                            except Exception:
                                img_bytes = None
                    if img_bytes:
                        image_present = True
                        fmt = 'png'
                        if img_bytes[:3] == b'\xff\xd8\xff':
                            fmt = 'jpeg'
                        if img_bytes[:8] == b'\x89PNG\r\n\x1a\n' or img_bytes[:3] == b'\xff\xd8\xff':
                            image_data = f"data:image/{fmt};base64," + base64.b64encode(img_bytes).decode('utf-8')

                all_questions.append({
                    'question_text': qtext,
                    'type': qtype,
                    'btl': btl,
                    'marks': marks,
                    'course_outcomes': co,
                    'course_outcomes_cell': co_raw,
                    'course_outcomes_numbers': co_multi_numbers,
                    'chapter': chapter,
                    'image': image_data,
                    'question_source_row': q_source_row,
                    'question_source_col': q_source_col,
                    'image_anchor_row': image_anchor_row,
                    'image_anchor_col': image_anchor_col,
                    'image_mapped_row': mapped_row,
                    'image_present': image_present,
                })

        if not all_questions:
            return {
                'questions': [],
                'warning': 'No questions parsed from specified CO sheets.'
            }
        return {'questions': all_questions}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Excel parse failed: {e}")


